import json
import datetime
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tests.appium.config.appium_config import AppiumConfig
from tests.e2e.utils.logger import get_logger

logger = get_logger("AppiumReportGenerator")

class AppiumTestResultItem:
    def __init__(self, name: str, suite: str, status: str, duration: float, error: str = "", screenshot: str = ""):
        self.name = name
        self.suite = suite
        self.status = status  # PASSED, FAILED, SKIPPED
        self.duration = duration
        self.error = error
        self.screenshot = screenshot
        self.timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

class AppiumReportGenerator:
    def __init__(self, results: List[AppiumTestResultItem]):
        self.results = results
        self.total = len(results)
        self.passed = sum(1 for r in results if r.status.upper() == "PASSED")
        self.failed = sum(1 for r in results if r.status.upper() == "FAILED")
        self.skipped = sum(1 for r in results if r.status.upper() == "SKIPPED")
        self.pass_rate = round((self.passed / self.total * 100), 2) if self.total > 0 else 0.0
        self.total_duration = round(sum(r.duration for r in results), 2)

    def generate_all(self):
        # Base Appium Output Directory
        project_root = Path(__file__).resolve().parent.parent.parent.parent
        appium_results_dir = project_root / "Test Results" / "Appium"
        excel_dir = appium_results_dir / "Excel"
        html_dir = appium_results_dir / "HTML"
        ss_dir = appium_results_dir / "Screenshots"
        logs_dir = appium_results_dir / "Logs"
        summary_dir = appium_results_dir / "Summary"

        for p in [appium_results_dir, excel_dir, html_dir, ss_dir, logs_dir, summary_dir]:
            p.mkdir(parents=True, exist_ok=True)

        self.generate_excel_report(excel_dir / "Automation_Appium_Test_Report.xlsx")
        self.generate_html_report(html_dir / "appium-execution-report.html")
        self.generate_markdown_summary(summary_dir / "summary.md")

    def generate_excel_report(self, excel_path: Path):
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # Sheet 1: Mobile Execution Summary
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Mobile Execution Summary"
        ws_summary.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=16, bold=True, color="102A63")
        section_font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10, color="0F172A")
        
        fill_navy = PatternFill(start_color="17212B", end_color="17212B", fill_type="solid")
        fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = "CephGrow AI - Android Appium E2E Automation Summary"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = Alignment(vertical="center")
        
        ws_summary["A3"] = "Target Application:"
        ws_summary["B3"] = f"{AppiumConfig.APP_PACKAGE} ({AppiumConfig.PLATFORM_NAME})"
        ws_summary["A4"] = "Execution Date:"
        ws_summary["B4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A3"].font = section_font
        ws_summary["A4"].font = section_font
        
        headers_summary = ["Metric", "Value", "Percentage", "Notes"]
        for col_idx, text in enumerate(headers_summary, start=1):
            cell = ws_summary.cell(row=6, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        metrics_data = [
            ("Total Mobile Test Cases", self.total, "100%", "Executed Appium Android E2E Scenarios"),
            ("Passed", self.passed, f"{self.pass_rate}%", "Verified on Android Emulator/Device"),
            ("Failed", self.failed, f"{round(self.failed/self.total*100, 2) if self.total else 0}%", "Requires investigation"),
            ("Skipped", self.skipped, f"{round(self.skipped/self.total*100, 2) if self.total else 0}%", "Skipped scenarios"),
            ("Total Execution Time", f"{self.total_duration}s", "-", "Combined Appium suite duration")
        ]
        
        for row_idx, row_data in enumerate(metrics_data, start=7):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx == 2 and row_data[0] == "Passed":
                    cell.fill = fill_pass
                elif col_idx == 2 and row_data[0] == "Failed" and self.failed > 0:
                    cell.fill = fill_fail

        # -------------------------------------------------------------
        # Sheet 2: Mobile Test Details
        # -------------------------------------------------------------
        ws_details = wb.create_sheet(title="Mobile Test Execution Details")
        ws_details.views.sheetView[0].showGridLines = True
        
        detail_headers = ["#", "Mobile Suite", "Appium Test Case Name", "Status", "Duration (s)", "Timestamp", "Failure Reason", "Screenshot Path"]
        for col_idx, text in enumerate(detail_headers, start=1):
            cell = ws_details.cell(row=1, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx, res in enumerate(self.results, start=2):
            ws_details.cell(row=row_idx, column=1, value=row_idx - 1).font = cell_font
            ws_details.cell(row=row_idx, column=2, value=res.suite).font = cell_font
            ws_details.cell(row=row_idx, column=3, value=res.name).font = cell_font
            
            status_cell = ws_details.cell(row=row_idx, column=4, value=res.status)
            status_cell.font = Font(name="Segoe UI", size=10, bold=True)
            status_cell.alignment = Alignment(horizontal="center")
            if res.status.upper() == "PASSED":
                status_cell.fill = fill_pass
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="15803D")
            elif res.status.upper() == "FAILED":
                status_cell.fill = fill_fail
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
                
            ws_details.cell(row=row_idx, column=5, value=f"{res.duration:.2f}").font = cell_font
            ws_details.cell(row=row_idx, column=6, value=res.timestamp).font = cell_font
            ws_details.cell(row=row_idx, column=7, value=res.error if res.error else "N/A").font = cell_font
            ws_details.cell(row=row_idx, column=8, value=res.screenshot if res.screenshot else "N/A").font = cell_font
            
            for col_idx in range(1, 9):
                ws_details.cell(row=row_idx, column=col_idx).border = thin_border
                
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
                
        wb.save(excel_path)
        logger.info(f"Appium Excel report saved successfully to {excel_path}")

    def generate_html_report(self, html_path: Path):
        rows_html = ""
        for idx, res in enumerate(self.results, 1):
            badge_class = "badge-pass" if res.status.upper() == "PASSED" else "badge-fail"
            error_msg = f'<div class="error-msg">{res.error}</div>' if res.error else ''
            screenshot_link = f'<a href="../Screenshots/{Path(res.screenshot).name}" target="_blank" class="ss-link">View Screenshot</a>' if res.screenshot else 'N/A'
            
            rows_html += f"""
            <tr>
                <td>{idx}</td>
                <td><strong>{res.suite}</strong></td>
                <td>{res.name}</td>
                <td><span class="badge {badge_class}">{res.status}</span></td>
                <td>{res.duration:.2f}s</td>
                <td>{screenshot_link}</td>
            </tr>
            {f'<tr><td colspan="6">{error_msg}</td></tr>' if error_msg else ''}
            """
            
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CephGrow AI - Android Appium Execution Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #F8FAFC; color: #0F172A; margin: 0; padding: 24px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .header {{ background: linear-gradient(135deg, #17212B 0%, #102A63 100%); color: white; padding: 28px; border-radius: 16px; margin-bottom: 24px; }}
        .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 16px; margin-bottom: 24px; }}
        .card {{ background: white; padding: 20px; border-radius: 12px; border: 1px solid #E2E8F0; }}
        .card-val {{ font-size: 32px; font-weight: 900; margin-top: 8px; }}
        .card-lbl {{ font-size: 12px; font-weight: 700; text-transform: uppercase; color: #64748B; }}
        table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 12px; overflow: hidden; border: 1px solid #E2E8F0; }}
        th, td {{ padding: 14px 18px; text-align: left; border-bottom: 1px solid #E2E8F0; font-size: 14px; }}
        th {{ background: #F1F5F9; font-weight: 800; color: #102A63; }}
        .badge {{ padding: 4px 10px; border-radius: 20px; font-size: 12px; font-weight: 800; }}
        .badge-pass {{ background: #DCFCE7; color: #15803D; }}
        .badge-fail {{ background: #FEE2E2; color: #B91C1C; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>CephGrow AI - Android Appium Automation Report</h1>
            <p>Target App: {AppiumConfig.APP_PACKAGE} | Platform: {AppiumConfig.PLATFORM_NAME} | Date: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
        <div class="grid">
            <div class="card"><div class="card-lbl">Total Mobile Tests</div><div class="card-val">{self.total}</div></div>
            <div class="card"><div class="card-lbl">Passed</div><div class="card-val" style="color:#16A34A">{self.passed}</div></div>
            <div class="card"><div class="card-lbl">Failed</div><div class="card-val" style="color:#DC2626">{self.failed}</div></div>
            <div class="card"><div class="card-lbl">Pass Rate</div><div class="card-val" style="color:#16A34A">{self.pass_rate}%</div></div>
        </div>
        <table>
            <thead>
                <tr><th>#</th><th>Suite</th><th>Appium Test Case</th><th>Status</th><th>Duration</th><th>Artifacts</th></tr>
            </thead>
            <tbody>{rows_html}</tbody>
        </table>
    </div>
</body>
</html>
"""
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        logger.info(f"Appium HTML report saved successfully to {html_path}")

    def generate_markdown_summary(self, summary_path: Path):
        failed_items = [r for r in self.results if r.status.upper() == "FAILED"]
        failed_text = ""
        if failed_items:
            for item in failed_items:
                failed_text += f"- **{item.name}** (`{item.suite}`)\n  - Reason: {item.error or 'Assertion error'}\n"
        else:
            failed_text = "None! All mobile Appium tests passed cleanly.\n"
            
        md_content = f"""# Android Appium Mobile Automation Summary

Target Package: `{AppiumConfig.APP_PACKAGE}`

Total Mobile Tests: {self.total}
Passed: {self.passed}
Failed: {self.failed}
Skipped: {self.skipped}
Pass Percentage: {self.pass_rate}%

Failed Mobile Tests:
{failed_text}
"""
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(md_content)
        logger.info(f"Appium markdown summary saved successfully to {summary_path}")
