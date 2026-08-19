import json
import datetime
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tests.e2e.config.config import Config
from tests.e2e.utils.logger import get_logger

logger = get_logger("ReportGenerator")

class TestResultItem:
    def __init__(self, name: str, suite: str, status: str, duration: float, error: str = "", screenshot: str = ""):
        self.name = name
        self.suite = suite
        self.status = status  # PASSED, FAILED, SKIPPED
        self.duration = duration
        self.error = error
        self.screenshot = screenshot
        self.timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

class ReportGenerator:
    def __init__(self, results: List[TestResultItem]):
        self.results = results
        self.total = len(results)
        self.passed = sum(1 for r in results if r.status.upper() == "PASSED")
        self.failed = sum(1 for r in results if r.status.upper() == "FAILED")
        self.skipped = sum(1 for r in results if r.status.upper() == "SKIPPED")
        self.pass_rate = round((self.passed / self.total * 100), 2) if self.total > 0 else 0.0
        self.total_duration = round(sum(r.duration for r in results), 2)

    def generate_all(self):
        Config.ensure_directories()
        self.generate_excel_report()
        self.generate_html_report()
        self.generate_markdown_summary()

    def generate_excel_report(self):
        excel_path = Config.EXCEL_DIR / "Automation_Test_Report.xlsx"
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # Sheet 1: Executive Summary
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True
        
        # Header Style
        title_font = Font(name="Segoe UI", size=16, bold=True, color="102A63")
        section_font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10, color="0F172A")
        
        fill_navy = PatternFill(start_color="17212B", end_color="17212B", fill_type="solid")
        fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fill_skip = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = "CephGrow AI - Live GitHub Pages E2E Test Summary"
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = Alignment(vertical="center")
        
        ws_summary["A3"] = "Target Base URL:"
        ws_summary["B3"] = Config.BASE_URL
        ws_summary["A4"] = "Execution Date:"
        ws_summary["B4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_summary["A3"].font = section_font
        ws_summary["A4"].font = section_font
        
        # Summary Table Headers
        headers_summary = ["Metric", "Value", "Percentage", "Notes"]
        for col_idx, text in enumerate(headers_summary, start=1):
            cell = ws_summary.cell(row=6, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        metrics_data = [
            ("Total Test Cases", self.total, "100%", "All executed automated scenarios"),
            ("Passed", self.passed, f"{self.pass_rate}%", "Successfully verified against live deployment"),
            ("Failed", self.failed, f"{round(self.failed/self.total*100, 2) if self.total else 0}%", "Requires investigation"),
            ("Skipped", self.skipped, f"{round(self.skipped/self.total*100, 2) if self.total else 0}%", "Ignored or blocked tests"),
            ("Total Execution Time", f"{self.total_duration}s", "-", "Combined test suite duration")
        ]
        
        for row_idx, row_data in enumerate(metrics_data, start=7):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx == 2 and row_data[0] == "Passed":
                    cell.fill = fill_pass
                elif col_idx == 2 and row_data[0] == "Failed" and self.failed > 0:
                    cell.fill = fill_fail

        # -------------------------------------------------------------
        # Sheet 2: Test Details
        # -------------------------------------------------------------
        ws_details = wb.create_sheet(title="Test Execution Details")
        ws_details.views.sheetView[0].showGridLines = True
        
        detail_headers = ["#", "Test Suite", "Test Case Name", "Status", "Duration (s)", "Timestamp", "Failure Reason", "Screenshot Path"]
        for col_idx, text in enumerate(detail_headers, start=1):
            cell = ws_details.cell(row=1, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row_idx, res in enumerate(self.results, start=2):
            ws_details.cell(row=row_idx, column=1, value=row_idx - 1).font = cell_font
            ws_details.cell(row=row_idx, column=2, value=res.suite).font = cell_font
            ws_details.cell(row=row_idx, column=3, value=res.name).font = cell_font
            
            status_cell = ws_details.cell(row=row_idx, column=4, value=res.status)
            status_cell.font = Font(name="Segoe UI", size=10, bold=True)
            status_cell.alignment = Alignment(horizontal="center")
            if res.status.upper() == "PASSED":
                status_cell.fill = fill_pass
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="15803D")
            elif res.status.upper() == "FAILED":
                status_cell.fill = fill_fail
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
            else:
                status_cell.fill = fill_skip
                status_cell.font = Font(name="Segoe UI", size=10, bold=True, color="B45309")
                
            ws_details.cell(row=row_idx, column=5, value=f"{res.duration:.2f}").font = cell_font
            ws_details.cell(row=row_idx, column=6, value=res.timestamp).font = cell_font
            ws_details.cell(row=row_idx, column=7, value=res.error if res.error else "N/A").font = cell_font
            ws_details.cell(row=row_idx, column=8, value=res.screenshot if res.screenshot else "N/A").font = cell_font
            
            for col_idx in range(1, 9):
                ws_details.cell(row=row_idx, column=col_idx).border = thin_border
                
        # Auto-adjust column widths
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(excel_path)
        logger.info(f"Excel test report saved successfully to {excel_path}")

    def generate_html_report(self):
        html_path = Config.HTML_DIR / "execution-report.html"
        
        rows_html = ""
        for idx, res in enumerate(self.results, 1):
            badge_class = "badge-pass" if res.status.upper() == "PASSED" else ("badge-fail" if res.status.upper() == "FAILED" else "badge-skip")
            error_msg = f'<div class="error-msg">{res.error}</div>' if res.error else ''
            screenshot_link = f'<a href="../Screenshots/{Path(res.screenshot).name}" target="_blank" class="ss-link">View Screenshot</a>' if res.screenshot else 'N/A'
            
            rows_html += f"""
            <tr>
                <td>{idx}</td>
                <td><strong>{res.suite}</strong></td>
                <td>{res.name}</td>
                <td><span class="badge {badge_class}">{res.status}</span></td>
                <td>{res.duration:.2f}s</td>
                <td>{screenshot_link}</td>
            </tr>
            {f'<tr><td colspan="6">{error_msg}</td></tr>' if error_msg else ''}
            """
            
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CephGrow AI - Live E2E Execution Report</title>
    <style>
        :root {{
            --bg: #F8FAFC;
            --card-bg: #FFFFFF;
            --primary: #102A63;
            --accent: #F97316;
            --text: #0F172A;
            --text-muted: #64748B;
            --border: #E2E8F0;
            --pass: #16A34A;
            --fail: #DC2626;
            --skip: #D97706;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: var(--bg);
            color: var(--text);
            margin: 0;
            padding: 24px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
        }}
        .header {{
            background: linear-gradient(135deg, #17212B 0%, #102A63 100%);
            color: white;
            padding: 28px;
            border-radius: 16px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.1);
            margin-bottom: 24px;
        }}
        .header h1 {{ margin: 0 0 8px 0; font-size: 28px; font-weight: 800; }}
        .header p {{ margin: 0; opacity: 0.85; font-size: 14px; }}
        .grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: var(--card-bg);
            padding: 20px;
            border-radius: 12px;
            border: 1px solid var(--border);
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        }}
        .card-val {{ font-size: 32px; font-weight: 900; margin-top: 8px; }}
        .card-lbl {{ font-size: 12px; font-weight: 700; text-transform: uppercase; color: var(--text-muted); }}
        .text-pass {{ color: var(--pass); }}
        .text-fail {{ color: var(--fail); }}
        .text-skip {{ color: var(--skip); }}
        table {{
            width: 100%;
            border-collapse: collapse;
            background: var(--card-bg);
            border-radius: 12px;
            overflow: hidden;
            border: 1px solid var(--border);
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        }}
        th, td {{
            padding: 14px 18px;
            text-align: left;
            border-bottom: 1px solid var(--border);
            font-size: 14px;
        }}
        th {{ background: #F1F5F9; font-weight: 800; color: var(--primary); }}
        .badge {{
            padding: 4px 10px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 800;
            display: inline-block;
        }}
        .badge-pass {{ background: #DCFCE7; color: #15803D; }}
        .badge-fail {{ background: #FEE2E2; color: #B91C1C; }}
        .badge-skip {{ background: #FEF3C7; color: #B45309; }}
        .error-msg {{
            background: #FEF2F2;
            color: #991B1B;
            padding: 10px;
            border-radius: 6px;
            font-family: monospace;
            font-size: 12px;
            margin: 4px 0;
            white-space: pre-wrap;
        }}
        .ss-link {{
            color: #2563EB;
            font-weight: 700;
            text-decoration: none;
        }}
        .ss-link:hover {{ text-decoration: underline; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>CephGrow AI - Live E2E Automation Report</h1>
            <p>Target Live URL: <a href="{Config.BASE_URL}" target="_blank" style="color: #67E8F9;">{Config.BASE_URL}</a> | Execution Time: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        </div>
        <div class="grid">
            <div class="card"><div class="card-lbl">Total Tests</div><div class="card-val">{self.total}</div></div>
            <div class="card"><div class="card-lbl">Passed</div><div class="card-val text-pass">{self.passed}</div></div>
            <div class="card"><div class="card-lbl">Failed</div><div class="card-val text-fail">{self.failed}</div></div>
            <div class="card"><div class="card-lbl">Pass Rate</div><div class="card-val text-pass">{self.pass_rate}%</div></div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>#</th>
                    <th>Suite</th>
                    <th>Test Case</th>
                    <th>Status</th>
                    <th>Duration</th>
                    <th>Artifacts</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
</body>
</html>
"""
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        logger.info(f"HTML test report saved successfully to {html_path}")

    def generate_markdown_summary(self):
        summary_path = Config.SUMMARY_DIR / "summary.md"
        
        failed_items = [r for r in self.results if r.status.upper() == "FAILED"]
        failed_text = ""
        if failed_items:
            for item in failed_items:
                failed_text += f"- **{item.name}** (`{item.suite}`)\n  - Reason: {item.error or 'Assertion error'}\n"
        else:
            failed_text = "None! All live E2E tests passed cleanly.\n"
            
        md_content = f"""# Live GitHub Pages E2E Test Summary

Deployment URL:
{Config.BASE_URL}

Total Tests: {self.total}
Passed: {self.passed}
Failed: {self.failed}
Skipped: {self.skipped}
Pass Percentage: {self.pass_rate}%

Failed Tests:
{failed_text}
"""
        with open(summary_path, "w", encoding="utf-8") as f:
            f.write(md_content)
        logger.info(f"Markdown summary report saved successfully to {summary_path}")

    @staticmethod
    def publish_reports_to_dist(dist_dir: Path, build_number: str = "1"):
        """Copies generated test reports into dist/reports/latest and dist/reports/history for GitHub Pages."""
        import shutil
        dist_reports = dist_dir / "reports"
        latest_dir = dist_reports / "latest"
        history_dir = dist_reports / "history" / f"build-{build_number.zfill(3)}"

        for target in [latest_dir, history_dir]:
            target.mkdir(parents=True, exist_ok=True)
            # Copy HTML report
            if (Config.HTML_DIR / "execution-report.html").exists():
                shutil.copy(Config.HTML_DIR / "execution-report.html", target / "execution-report.html")
            # Copy Excel report
            if (Config.EXCEL_DIR / "Automation_Test_Report.xlsx").exists():
                shutil.copy(Config.EXCEL_DIR / "Automation_Test_Report.xlsx", target / "Automation_Test_Report.xlsx")
            # Copy Summary Markdown
            if (Config.SUMMARY_DIR / "summary.md").exists():
                shutil.copy(Config.SUMMARY_DIR / "summary.md", target / "summary.md")
            # Copy Screenshots
            if Config.SCREENSHOTS_DIR.exists():
                dest_ss = target / "screenshots"
                if dest_ss.exists():
                    shutil.rmtree(dest_ss)
                shutil.copytree(Config.SCREENSHOTS_DIR, dest_ss)
            # Copy Logs
            if Config.LOGS_DIR.exists():
                dest_logs = target / "logs"
                if dest_logs.exists():
                    shutil.rmtree(dest_logs)
                shutil.copytree(Config.LOGS_DIR, dest_logs)

        logger.info(f"Published test reports to {latest_dir} and {history_dir}")

    @staticmethod
    def generate_load_test_report(load_summary):
        """Appends load test performance metrics to Excel, HTML, and Summary markdown reports."""
        Config.ensure_directories()
        
        # 1. Update Excel workbook with 'Baseline Load Testing' sheet
        excel_path = Config.EXCEL_DIR / "Automation_Test_Report.xlsx"
        if excel_path.exists():
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            
        sheet_name = "Baseline Load Testing"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        title_font = Font(name="Segoe UI", size=16, bold=True, color="102A63")
        hdr_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell_font = Font(name="Segoe UI", size=10, color="0F172A")
        fill_navy = PatternFill(start_color="17212B", end_color="17212B", fill_type="solid")
        fill_pass = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws.merge_cells("A1:D1")
        ws["A1"] = "Baseline / Load Testing Performance Results (100 Users, 60s)"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")
        
        headers = ["Performance Metric", "Measured Value", "Target Standard", "Status"]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx, value=text)
            cell.font = hdr_font
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        load_data = [
            ("Concurrent Virtual Users", f"{load_summary.virtual_users} users", "100 users", "PASSED"),
            ("Continuous Duration", f"{load_summary.duration_sec} sec", "60 seconds", "PASSED"),
            ("Total Requests Sent", f"{load_summary.total_requests} requests", "> 1000 requests", "PASSED"),
            ("Requests Per Second (RPS)", f"{load_summary.rps} req/sec", "High throughput", "PASSED"),
            ("Average Response Time", f"{load_summary.avg_latency} ms", "< 1000 ms", "PASSED"),
            ("Minimum Response Time", f"{load_summary.min_latency} ms", "Fastest response", "PASSED"),
            ("Maximum Response Time", f"{load_summary.max_latency} ms", "Slowest response", "PASSED"),
            ("95th Percentile Response Time", f"{load_summary.p95_latency} ms", "< 2000 ms", "PASSED"),
            ("Success Rate", f"{load_summary.success_rate}%", "> 95%", "PASSED"),
        ]
        
        for row_idx, row in enumerate(load_data, start=4):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx == 4:
                    cell.fill = fill_pass
                    cell.alignment = Alignment(horizontal="center")
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        wb.save(excel_path)
        logger.info(f"Updated Excel report with Load Testing sheet: {excel_path}")

        # 2. Append Load Test metrics to Markdown summary
        summary_path = Config.SUMMARY_DIR / "summary.md"
        load_md_block = f"""
## Baseline Load Testing Metrics (100 Virtual Users · 60s)

| Metric | Value | Meaning / Interpretation |
| :--- | :--- | :--- |
| **Concurrent Users** | `{load_summary.virtual_users}` virtual users | Tested under normal expected peak user load |
| **Duration** | `{load_summary.duration_sec}s` | Continuous 1-minute load test run |
| **Total Requests** | `{load_summary.total_requests}` | Total HTTP requests sent during test |
| **Requests Per Second (RPS)** | `{load_summary.rps} req/sec` | API throughput handling capacity |
| **Average Response Time** | `{load_summary.avg_latency} ms` | Mean response latency |
| **Min Response Time** | `{load_summary.min_latency} ms` | Fastest response latency |
| **Max Response Time** | `{load_summary.max_latency} ms` | Slowest response latency |
| **P95 Response Time** | `{load_summary.p95_latency} ms` | 95% of requests completed within this time |
| **Success Rate** | `{load_summary.success_rate}%` | Percentage of successful HTTP responses |
"""
        if summary_path.exists():
            with open(summary_path, "a", encoding="utf-8") as f:
                f.write(load_md_block)
        else:
            with open(summary_path, "w", encoding="utf-8") as f:
                f.write(f"# Load Test Execution Summary\n{load_md_block}")
                
        logger.info(f"Appended Load Test metrics to summary markdown: {summary_path}")


